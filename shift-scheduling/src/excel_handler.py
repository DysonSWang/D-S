"""
Excel 导入导出处理
"""
from typing import List, Dict, Optional, Set
from datetime import date
from pathlib import Path
import json

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    from .models import (
        Employee, Schedule, ShiftAssignment, ShiftType,
        ShiftCycle, SchedulingConfig
    )
except ImportError:
    from models import (
        Employee, Schedule, ShiftAssignment, ShiftType,
        ShiftCycle, SchedulingConfig
    )


class ExcelHandler:
    """Excel 文件处理"""
    
    # 班次颜色
    SHIFT_COLORS = {
        ShiftType.G: "FFFFD700",  # 金色
        ShiftType.T: "87CEEB",    # 天蓝色
        ShiftType.Y: "9370DB",    # 紫色
        ShiftType.OFF: "E0E0E0",  # 灰色
    }
    
    def __init__(self):
        pass
    
    def export_employees_template(self, output_path: str) -> str:
        """
        导出员工信息模板
        
        列：员工 ID | 姓名 | 组别 | 固定休息天数 | 优先休息日期 (JSON)
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "员工信息"
        
        # 表头
        headers = ["员工 ID", "姓名", "组别", "固定休息天数", "优先休息日期 (YYYY-MM-DD 格式，多个用逗号分隔)"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # 示例数据
        example = [
            ["EMP0001", "张三", "组 01", 8, "2024-01-01,2024-01-15"],
            ["EMP0002", "李四", "组 01", 8, ""],
            ["EMP0003", "王五", "组 02", 8, "2024-01-10"],
        ]
        
        for row_idx, row_data in enumerate(example, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # 设置列宽
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 40
        
        # 保存
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(output_path))
        
        return str(output_path)
    
    def import_employees(self, input_path: str) -> List[Employee]:
        """
        导入员工信息
        
        Returns:
            List[Employee]: 员工列表
        """
        employees = []
        wb = load_workbook(input_path)
        ws = wb.active
        
        # 跳过表头
        for row_idx in range(2, ws.max_row + 1):
            emp_id = ws.cell(row=row_idx, column=1).value
            name = ws.cell(row=row_idx, column=2).value
            group = ws.cell(row=row_idx, column=3).value
            off_days = ws.cell(row=row_idx, column=4).value
            pref_dates_str = ws.cell(row=row_idx, column=5).value
            
            if not emp_id:
                continue
            
            # 解析优先休息日期
            preferred_dates = set()
            if pref_dates_str:
                for date_str in str(pref_dates_str).split(','):
                    date_str = date_str.strip()
                    if date_str:
                        try:
                            d = date.fromisoformat(date_str)
                            preferred_dates.add(d)
                        except ValueError:
                            pass
            
            emp = Employee(
                id=str(emp_id),
                name=str(name) if name else str(emp_id),
                group=str(group) if group else "默认组",
                fixed_off_days=int(off_days) if off_days else 8,
                preferred_off_dates=preferred_dates
            )
            employees.append(emp)
        
        return employees
    
    def export_schedule(
        self,
        schedule: Schedule,
        employees: List[Employee],
        output_path: str,
        group_by: str = "group"
    ) -> str:
        """
        导出排班结果
        
        Args:
            schedule: 排班结果
            employees: 员工列表
            output_path: 输出路径
            group_by: 分组方式 ("group" 或 "all")
        
        Returns:
            str: 输出文件路径
        """
        # 生成日期列表
        from datetime import timedelta
        dates = []
        current = schedule.cycle.start_date
        while current <= schedule.cycle.end_date:
            dates.append(current)
            current += timedelta(days=1)
        
        if group_by == "group":
            # 按组别分 sheet
            employees_by_group = {}
            for emp in employees:
                if emp.group not in employees_by_group:
                    employees_by_group[emp.group] = []
                employees_by_group[emp.group].append(emp)
            
            wb = Workbook()
            # 删除默认 sheet
            wb.remove(wb.active)
            
            for group_name, group_employees in sorted(employees_by_group.items()):
                ws = wb.create_sheet(title=f"{group_name}")
                self._write_schedule_sheet(ws, group_employees, dates, schedule)
        else:
            # 所有人在一个 sheet
            wb = Workbook()
            ws = wb.active
            ws.title = "排班表"
            self._write_schedule_sheet(ws, employees, dates, schedule)
        
        # 保存
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(output_path))
        
        return str(output_path)
    
    def _write_schedule_sheet(
        self,
        ws,
        employees: List[Employee],
        dates: List[date],
        schedule: Schedule
    ):
        """写入排班 sheet"""
        # 表头：员工信息 + 日期
        ws.cell(row=1, column=1, value="员工 ID").font = Font(bold=True)
        ws.cell(row=1, column=2, value="姓名").font = Font(bold=True)
        ws.cell(row=1, column=3, value="组别").font = Font(bold=True)
        
        for col_idx, d in enumerate(dates, 4):
            cell = ws.cell(row=1, column=col_idx, value=d.strftime("%m/%d"))
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            # 周末标记
            if d.weekday() >= 5:
                cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        
        # 设置列宽
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 8
        for col_idx in range(4, len(dates) + 4):
            ws.column_dimensions[get_column_letter(col_idx)].width = 5
        
        # 边框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 员工数据
        for row_idx, emp in enumerate(employees, 2):
            ws.cell(row=row_idx, column=1, value=emp.id).border = thin_border
            ws.cell(row=row_idx, column=2, value=emp.name).border = thin_border
            ws.cell(row=row_idx, column=3, value=emp.group).border = thin_border
            
            # 获取员工排班
            emp_shifts = {a.date: a.shift for a in schedule.assignments.get(emp.id, [])}
            
            for col_idx, d in enumerate(dates, 4):
                shift = emp_shifts.get(d, ShiftType.OFF)
                cell = ws.cell(row=row_idx, column=col_idx, value=shift.value)
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border
                
                # 设置颜色
                color = self.SHIFT_COLORS.get(shift, "FFFFFF")
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        
        # 底部统计行
        stat_row = len(employees) + 2
        ws.cell(row=stat_row, column=1, value="统计").font = Font(bold=True)
        
        for col_idx, d in enumerate(dates, 4):
            # 统计当天各班次人数
            counts = {ShiftType.G: 0, ShiftType.T: 0, ShiftType.Y: 0, ShiftType.OFF: 0}
            for emp in employees:
                emp_shifts = schedule.assignments.get(emp.id, [])
                for a in emp_shifts:
                    if a.date == d:
                        counts[a.shift] = counts.get(a.shift, 0) + 1
            
            summary = f"G:{counts[ShiftType.G]}/T:{counts[ShiftType.T]}/Y:{counts[ShiftType.Y]}"
            cell = ws.cell(row=stat_row, column=col_idx, value=summary)
            cell.font = Font(bold=True, size=8)
            cell.alignment = Alignment(horizontal="center")
    
    def export_config(self, config: SchedulingConfig, output_path: str) -> str:
        """导出配置文件"""
        config_dict = {
            "shift_ratio": {
                k.value: v for k, v in config.shift_ratio.items()
            },
            "off_days_per_month": config.off_days_per_month,
            "max_consecutive_work_days": config.max_consecutive_work_days,
            "consecutive_work_threshold": config.consecutive_work_threshold,
            "min_shift_types_after_threshold": config.min_shift_types_after_threshold,
            "total_employees": config.total_employees
        }
        
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(config_dict, f, ensure_ascii=False, indent=2)
        
        return str(output_path)
    
    def import_config(self, input_path: str) -> SchedulingConfig:
        """导入配置文件"""
        with open(input_path, 'r', encoding='utf-8') as f:
            config_dict = json.load(f)
        
        shift_ratio = {
            ShiftType(k): v for k, v in config_dict.get("shift_ratio", {}).items()
        }
        
        return SchedulingConfig(
            shift_ratio=shift_ratio,
            off_days_per_month=config_dict.get("off_days_per_month", 8),
            max_consecutive_work_days=config_dict.get("max_consecutive_work_days", 6),
            consecutive_work_threshold=config_dict.get("consecutive_work_threshold", 3),
            min_shift_types_after_threshold=config_dict.get("min_shift_types_after_threshold", 2),
            total_employees=config_dict.get("total_employees", 200)
        )
